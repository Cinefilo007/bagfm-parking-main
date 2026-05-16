import uuid
import io
import zipfile
import qrcode
from datetime import datetime, date, timedelta, timezone
from typing import List, Optional, Tuple
from sqlalchemy.future import select
from sqlalchemy import func, and_, or_, delete
from sqlalchemy.ext.asyncio import AsyncSession
from supabase import create_client, Client

from app.core.config import obtener_config
from app.models.alcabala_evento import LotePaseMasivo, SolicitudEvento
from app.models.codigo_qr import CodigoQR
from app.models.enums import PasseTipo, QRTipo, RolTipo
from app.models.usuario import Usuario
from app.models.vehiculo import Vehiculo
from app.models.asignacion_zona import AsignacionZona
from app.models.tipo_acceso_custom import TipoAccesoCustom
from app.models.entidad_civil import EntidadCivil
from app.core.security import crear_token_evento
# Importación diferida para evitar ciclos

config = obtener_config()

class PaseService:
    def __init__(self):
        try:
            if config.supabase_url and config.supabase_service_key and config.supabase_url.startswith("http"):
                self.supabase: Client = create_client(config.supabase_url, config.supabase_service_key)
            else:
                self.supabase = None
        except Exception as e:
            print(f"ALERTA TÁCTICA: Fallo al inicializar Supabase Storage: {e}")
            self.supabase = None

    async def crear_lote(
        self, 
        db: AsyncSession, 
        datos: dict, 
        creado_por_id: uuid.UUID,
        solicitud_id: Optional[uuid.UUID] = None
    ) -> LotePaseMasivo:
        """
        Crea un lote de pases masivos con verificación de cuota y capacidad de estacionamiento.
        Sigue la Lógica Táctica v2.2: 
        - General: usa puestos no reservados.
        - Staff/VIP/etc: usa puestos reservados para su categoría.
        """
        # 1. Verificar Cuota y Entidad
        usuario = await db.get(Usuario, creado_por_id)
        if not usuario or not usuario.entidad_id:
             # Si no tiene entidad, solo permitimos si es ADMIN_BASE o COMANDANTE (sin cuota por ahora)
             if usuario and usuario.rol not in [RolTipo.COMANDANTE, RolTipo.ADMIN_BASE]:
                raise ValueError("El usuario no tiene una entidad vinculada.")
        
        entidad = None
        if usuario.entidad_id:
            entidad = await db.get(EntidadCivil, usuario.entidad_id)
            if not entidad:
                raise ValueError("Entidad no encontrada.")
            
            # La capacidad se valida en el paso 2 (VALIDACIÓN TÁCTICA DE ESTACIONAMIENTO)
            # a través de las AsignacionZona. La cuota autónoma legacy ya no se aplica.

        # 2. VALIDACIÓN TÁCTICA DE ESTACIONAMIENTO (Requerimiento #8)
        zona_id = datos.get('zona_asignada_id') or datos.get('zona_id')
        tipo_acc = datos.get('tipo_acceso', 'general')
        
        # Salvaguarda: si tipo_acceso='custom' pero sin ID, tratar como 'general'
        custom_id = datos.get('tipo_acceso_custom_id')
        if tipo_acc == 'custom' and not custom_id:
            tipo_acc = 'general'

        plan_distribucion = None
        if entidad:
            # 2a. Parsear fechas
            try:
                fecha_ini = datetime.strptime(datos['fecha_inicio'], '%Y-%m-%d').date() if isinstance(datos['fecha_inicio'], str) else datos['fecha_inicio']
                fecha_fin = datetime.strptime(datos['fecha_fin'], '%Y-%m-%d').date() if isinstance(datos['fecha_fin'], str) else datos['fecha_fin']
            except Exception:
                fecha_ini = date.today()
                fecha_fin = date.today()

            # 2b. Validación por Zona Específica (si se seleccionó una)
            if zona_id:
                query_asig = select(AsignacionZona).where(
                    and_(
                        AsignacionZona.entidad_id == entidad.id, 
                        AsignacionZona.zona_id == zona_id,
                        AsignacionZona.activa == True
                    )
                )
                asig_especifica = (await db.execute(query_asig)).scalar()
                
                if asig_especifica:
                    ocupacion_actual = await self.calcular_ocupacion_proyectada(db, zona_id, fecha_ini, fecha_fin)
                    cupo_total = asig_especifica.cupo_asignado
                    cupo_disponible = max(0, cupo_total - ocupacion_actual)
                    
                    dist_auto = datos.get('distribucion_automatica') or datos.get('distribucion_automatic', False)
                    if datos['cantidad_pases'] > cupo_disponible:
                        if not dist_auto:
                            raise ValueError(f"CAPACIDAD AGOTADA para estas fechas: Libre {cupo_disponible}, Requerido {datos['cantidad_pases']}.")
                        else:
                            # Auto-distribución desde una zona principal
                            res_plan = await self.sugerir_distribucion_entidad(db, entidad.id, datos['cantidad_pases'], fecha_ini, fecha_fin)
                            if not res_plan["completo"]:
                                raise ValueError(f"CAPACIDAD TOTAL SUPERADA: La entidad solo tiene {datos['cantidad_pases'] - res_plan['cantidad_restante']} puestos disponibles en total.")
                            
                            plan_distribucion = res_plan["distribucion"]
                            # Priorizar la zona elegida poniéndola de primera en el plan
                            plan_distribucion.sort(key=lambda x: 0 if str(x["zona_id"]) == str(zona_id) else 1)
            else:
                # 2c. DISTRIBUCIÓN AUTOMÁTICA (Si no hay zona_id)
                res_plan = await self.sugerir_distribucion_entidad(db, entidad.id, datos['cantidad_pases'], fecha_ini, fecha_fin)
                if not res_plan["completo"]:
                    # Si no hay espacio suficiente en todas las zonas juntas
                    raise ValueError(f"CAPACIDAD TOTAL SUPERADA: La entidad solo tiene {datos['cantidad_pases'] - res_plan['cantidad_restante']} puestos disponibles en total.")
                
                plan_distribucion = res_plan["distribucion"]
                # Usar la primera zona lograda como referencia para el lote
                if plan_distribucion:
                    zona_id = plan_distribucion[0]["zona_id"]

        # 3. Generar serial y persistir lote
        serial_lote = await self._generar_serial_lote(db)
        
        nuevo_lote = LotePaseMasivo(
            codigo_serial=serial_lote,
            nombre_evento=datos['nombre_evento'],
            tipo_pase=datos['tipo_pase'],
            fecha_inicio=datos['fecha_inicio'],
            fecha_fin=datos['fecha_fin'],
            cantidad_pases=datos['cantidad_pases'],
            max_accesos_por_pase=datos.get('max_accesos_por_pase'),
            entidad_id=entidad.id if entidad else None,
            tipo_acceso=tipo_acc,
            tipo_acceso_custom_id=datos.get('tipo_acceso_custom_id'),
            zona_estacionamiento_id=zona_id,
            multi_vehiculo=datos.get('multi_vehiculo', False),
            max_vehiculos=datos.get('max_vehiculos', 1),
            creado_por=creado_por_id
        )
        db.add(nuevo_lote)
        await db.flush()

        # 4. Generar QRs con plan de distribución
        tipo_lote_val = nuevo_lote.tipo_pase.value if hasattr(nuevo_lote.tipo_pase, 'value') else str(nuevo_lote.tipo_pase)
        
        # Inyectar plan en extras para los generadores
        extras = {**datos, "plan_distribucion": plan_distribucion}

        if tipo_lote_val == PasseTipo.simple.value:
            await self._generar_pases_simples(db, nuevo_lote, creado_por_id, extras)
        elif tipo_lote_val == PasseTipo.portal.value:
            if datos.get('excel_data'):
                # Caso Híbrido: Autoregistro con datos parciales de Excel
                await self.procesar_json_identificado(db, nuevo_lote, datos['excel_data'], creado_por_id, extras)
            else:
                # Caso Estándar: Autoregistro vacío
                await self._generar_pases_portal(db, nuevo_lote, creado_por_id, extras)
        elif tipo_lote_val == PasseTipo.identificado.value:
            if datos.get('excel_data'):
                await self.procesar_json_identificado(db, nuevo_lote, datos['excel_data'], creado_por_id, extras)
            else:
                # Si es identificado pero no hay excel, generamos pases portal para que los llenen
                await self._generar_pases_portal(db, nuevo_lote, creado_por_id, extras)
        
        if solicitud_id:
            solicitud = await db.get(SolicitudEvento, solicitud_id)
            if solicitud:
                solicitud.lote_id = nuevo_lote.id

        await db.commit()
        await db.refresh(nuevo_lote, attribute_names=["zona_asignada", "tipo_acceso_custom"])
        return nuevo_lote

    async def _generar_serial_lote(self, db: AsyncSession) -> str:
        """Genera serial corto tipo B26401 (B+YY+M+CONSECUTIVO) verificado contra DB."""
        ahora = datetime.now()
        meses_id = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "A", "B", "C"]
        prefijo = f"B{str(ahora.year)[2:]}{meses_id[ahora.month-1]}"
        
        inicio_mes = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        query_count = select(func.count(LotePaseMasivo.id)).where(LotePaseMasivo.created_at >= inicio_mes)
        count = (await db.execute(query_count)).scalar() or 0
        
        tentativo = count + 1
        while True:
            serial = f"{prefijo}{str(tentativo).zfill(2)}"
            # Verificación física para evitar colisiones por borrados previos
            q_exists = select(LotePaseMasivo.id).where(LotePaseMasivo.codigo_serial == serial)
            if not (await db.execute(q_exists)).first():
                return serial
            tentativo += 1

    async def calcular_ocupacion_proyectada(self, db: AsyncSession, zona_id: uuid.UUID, inicio: date, fin: date) -> int:
        """Suma pases individuales que se traslapan con el periodo solicitado basándose en los QRs asignados."""
        from sqlalchemy import and_
        query = (
            select(func.count(CodigoQR.id))
            .join(LotePaseMasivo, CodigoQR.lote_id == LotePaseMasivo.id)
            .where(
                and_(
                    CodigoQR.zona_asignada_id == zona_id,
                    CodigoQR.activo == True,
                    LotePaseMasivo.fecha_inicio <= fin,
                    LotePaseMasivo.fecha_fin >= inicio
                )
            )
        )
        res = await db.execute(query)
        return int(res.scalar() or 0)

    async def sugerir_distribucion_entidad(self, db: AsyncSession, entidad_id: uuid.UUID, cantidad: int, inicio: date, fin: date):
        """Busca todas las asignaciones de la entidad y reparte la cantidad según disponibilidad temporal."""
        from app.models.asignacion_zona import AsignacionZona
        
        # 1. Obtener todas las asignaciones activas
        query = select(AsignacionZona).where(
            AsignacionZona.entidad_id == entidad_id,
            AsignacionZona.activa == True
        )
        res_asigs = await db.execute(query)
        asigs = res_asigs.scalars().all()
        
        resultados = []
        restante = cantidad
        
        # 2. Analizar disponibilidad real por zona considerando traslapes
        for asig in asigs:
            if restante <= 0: break
            
            ocupacion = await self.calcular_ocupacion_proyectada(db, asig.zona_id, inicio, fin)
            disponible = max(0, asig.cupo_asignado - ocupacion)
            
            if disponible > 0:
                tomar = min(disponible, restante)
                resultados.append({
                    "zona_id": asig.zona_id,
                    "zona_nombre": asig.zona_nombre,
                    "cupo_libre": disponible,
                    "sugerencia": tomar
                })
                restante -= tomar
        
        return {
            "distribucion": resultados,
            "cantidad_restante": restante,
            "completo": restante == 0
        }

    async def contar_ocupados_en_zona_por_entidad(self, db: AsyncSession, zona_id: uuid.UUID, entidad_id: uuid.UUID) -> int:
        """
        Cuenta los VehiculoPase (autos físicamente estacionados / ingresados) 
        en la zona que pertenecen a esta entidad (a través de lotes de pases).
        """
        from sqlalchemy import and_
        from app.models.vehiculo_pase import VehiculoPase
        
        query = (
            select(func.count(VehiculoPase.id))
            .join(CodigoQR, VehiculoPase.qr_id == CodigoQR.id)
            .join(LotePaseMasivo, CodigoQR.lote_id == LotePaseMasivo.id)
            .where(
                and_(
                    VehiculoPase.zona_asignada_id == zona_id,
                    VehiculoPase.ingresado == True,
                    LotePaseMasivo.entidad_id == entidad_id
                )
            )
        )
        return int((await db.execute(query)).scalar() or 0)

    async def contar_pases_activos_en_zona_para_fecha(
        self, db: AsyncSession, zona_id: uuid.UUID, fecha: date, limite: int = 20, offset: int = 0, busqueda: str = None
    ) -> dict:
        """
        Devuelve el conteo de QRs vigentes en una zona para un día específico
        y una muestra con paginación para el panel compacto del frontend.
        Se agregó parámetro de búsqueda.
        """
        from sqlalchemy import and_, or_

        condiciones = [
            CodigoQR.zona_asignada_id == zona_id,
            CodigoQR.activo == True,
            LotePaseMasivo.fecha_inicio <= fecha,
            LotePaseMasivo.fecha_fin >= fecha,
        ]

        if busqueda:
            term = f"%{busqueda.lower()}%"
            condiciones.append(or_(
                func.lower(CodigoQR.serial_legible).like(term),
                func.lower(CodigoQR.nombre_portador).like(term),
                func.lower(CodigoQR.cedula_portador).like(term),
                func.lower(CodigoQR.vehiculo_placa).like(term),
                func.lower(CodigoQR.vehiculo_modelo).like(term),
                func.lower(CodigoQR.vehiculo_marca).like(term),
            ))

        _WHERE = and_(*condiciones)

        # Total
        q_count = (
            select(func.count(CodigoQR.id))
            .join(LotePaseMasivo, CodigoQR.lote_id == LotePaseMasivo.id)
            .where(_WHERE)
        )
        total = int((await db.execute(q_count)).scalar() or 0)

        # Muestra paginada con campos enriquecidos
        q_muestra = (
            select(
                CodigoQR.id,
                CodigoQR.serial_legible,
                CodigoQR.nombre_portador,
                CodigoQR.cedula_portador,
                CodigoQR.vehiculo_placa,
                CodigoQR.datos_completos,
                CodigoQR.puesto_asignado_id,
                CodigoQR.tipo_acceso,
                CodigoQR.tipo_acceso_custom_id,
                CodigoQR.vehiculo_marca,
                CodigoQR.vehiculo_modelo,
                CodigoQR.vehiculo_color,
                CodigoQR.email_portador,
                CodigoQR.telefono_portador,
            )
            .join(LotePaseMasivo, CodigoQR.lote_id == LotePaseMasivo.id)
            .where(_WHERE)
            .order_by(CodigoQR.serial_legible.asc())
            .offset(offset)
            .limit(limite)
        )
        rows = (await db.execute(q_muestra)).all()

        muestra = [
            {
                "id": str(r.id),
                "serial_legible": r.serial_legible,
                "nombre_portador": r.nombre_portador,
                "cedula_portador": r.cedula_portador,
                "vehiculo_placa": r.vehiculo_placa,
                "vehiculo_marca": r.vehiculo_marca,
                "vehiculo_modelo": r.vehiculo_modelo,
                "vehiculo_color": r.vehiculo_color,
                "email_portador": r.email_portador,
                "telefono_portador": r.telefono_portador,
                "tiene_datos": bool(r.datos_completos or r.nombre_portador),
                "puesto_asignado_id": str(r.puesto_asignado_id) if r.puesto_asignado_id else None,
                "tipo_acceso": r.tipo_acceso.value if r.tipo_acceso else "general",
                "tipo_acceso_custom_id": str(r.tipo_acceso_custom_id) if r.tipo_acceso_custom_id else None,
            }
            for r in rows
        ]

        return {"total": total, "pases": muestra}

    async def obtener_fechas_con_lotes_por_entidad(
        self, db: AsyncSession, entidad_id: uuid.UUID
    ) -> dict:
        """
        Retorna un dict { 'YYYY-MM-DD': ['Zona A', 'Zona B'] } con todas las fechas
        en las que la entidad tiene lotes activos (no vencidos).
        Usado por el CalendarioLotes en el frontend para colorear días.
        """

        from app.models.zona_estacionamiento import ZonaEstacionamiento

        hoy = date.today()

        # Traer lotes activos (no vencidos) de la entidad
        q = (
            select(
                LotePaseMasivo.fecha_inicio,
                LotePaseMasivo.fecha_fin,
                LotePaseMasivo.zona_estacionamiento_id,
            )
            .where(
                and_(
                    LotePaseMasivo.entidad_id == entidad_id,
                    LotePaseMasivo.fecha_fin >= hoy,
                )
            )
        )
        lotes = (await db.execute(q)).all()

        # Resolver nombres de zona en una sola query
        zona_ids = list({l.zona_estacionamiento_id for l in lotes if l.zona_estacionamiento_id})
        zonas_map = {}
        if zona_ids:
            q_zonas = select(ZonaEstacionamiento.id, ZonaEstacionamiento.nombre).where(
                ZonaEstacionamiento.id.in_(zona_ids)
            )
            for row in (await db.execute(q_zonas)).all():
                zonas_map[str(row.id)] = row.nombre

        # Construir el dict de fechas → zonas
        calendario: dict = {}
        for lote in lotes:
            if not lote.fecha_inicio or not lote.fecha_fin:
                continue
            zona_nombre = zonas_map.get(str(lote.zona_estacionamiento_id), "Sin zona")
            # Expandir cada día del rango del lote
            dia = lote.fecha_inicio if isinstance(lote.fecha_inicio, date) else lote.fecha_inicio.date()
            fin = lote.fecha_fin if isinstance(lote.fecha_fin, date) else lote.fecha_fin.date()
            while dia <= fin:
                key = dia.isoformat()
                if key not in calendario:
                    calendario[key] = []
                if zona_nombre not in calendario[key]:
                    calendario[key].append(zona_nombre)
                dia += timedelta(days=1)

        return calendario


    async def calcular_ocupacion_por_tipo(self, db: AsyncSession, zona_id: uuid.UUID, tipo_acceso: str, tipo_acceso_custom_id: uuid.UUID, inicio: date, fin: date) -> int:
        """Calcula pases exactos del mismo tipo de acceso que se traslapan con el periodo en la zona basándose en los QRs asignados."""
        from sqlalchemy import and_
        
        conditions = [
            CodigoQR.zona_asignada_id == zona_id,
            CodigoQR.activo == True,
            LotePaseMasivo.fecha_inicio <= fin,
            LotePaseMasivo.fecha_fin >= inicio,
        ]
        
        if tipo_acceso == 'custom' and tipo_acceso_custom_id:
            conditions.append(CodigoQR.tipo_acceso_custom_id == tipo_acceso_custom_id)
        else:
            conditions.append(CodigoQR.tipo_acceso == tipo_acceso)
        
        query = (
            select(func.count(CodigoQR.id))
            .join(LotePaseMasivo, CodigoQR.lote_id == LotePaseMasivo.id)
            .where(and_(*conditions))
        )
        res = await db.execute(query)
        return int(res.scalar() or 0)

    async def validar_capacidad_completa(
        self, db: AsyncSession, entidad_id: uuid.UUID,
        zona_id: uuid.UUID, tipo_acceso: str, tipo_acceso_custom_id: uuid.UUID,
        cantidad: int, inicio: date, fin: date
    ) -> dict:
        """
        Validación Inteligente de Capacidad v3.0 — 3 Niveles con sugerencias contextuales.
        
        Nivel 1: Validación por categoría → Alerta + sugerencia de redistribución
        Nivel 2: Validación por zona → Alerta + sugerencia de usar otras zonas
        Nivel 3: Validación total entidad → Bloqueo duro
        """

        
        alertas = []
        sugerencias = []
        puede_crear = True
        
        # ── Datos base de la entidad ──────────────────────────────────────
        query_asigs = select(AsignacionZona).where(
            and_(AsignacionZona.entidad_id == entidad_id, AsignacionZona.activa == True)
        )
        asigs = (await db.execute(query_asigs)).scalars().all()
        
        capacidad_total_entidad = sum(a.cupo_asignado for a in asigs)
        ocupacion_total_entidad = 0
        for a in asigs:
            ocupacion_total_entidad += await self.calcular_ocupacion_proyectada(db, a.zona_id, inicio, fin)
        
        disponible_total_entidad = max(0, capacidad_total_entidad - ocupacion_total_entidad)
        
        # Resolver nombre del tipo de acceso
        nombre_tipo = tipo_acceso
        if tipo_acceso == 'custom' and tipo_acceso_custom_id:
            tc = await db.get(TipoAccesoCustom, tipo_acceso_custom_id)
            nombre_tipo = tc.nombre if tc else 'Personalizado'
        
        # ── NIVEL 1: Validación por Categoría ─────────────────────────────
        cupo_categoria = None
        ocupacion_categoria = 0
        disponible_categoria = None
        
        if zona_id:
            asig_zona = next((a for a in asigs if str(a.zona_id) == str(zona_id)), None)
            
            if asig_zona and asig_zona.distribucion_cupos:
                dist = asig_zona.distribucion_cupos
                # Buscar cupo para este tipo (por nombre o UUID)
                cupo_key = None
                for k in dist.keys():
                    if tipo_acceso == 'custom' and tipo_acceso_custom_id:
                        if k == str(tipo_acceso_custom_id) or k.lower() == nombre_tipo.lower():
                            cupo_key = k
                            break
                    else:
                        if k.lower() == tipo_acceso.lower() or k.lower() == nombre_tipo.lower():
                            cupo_key = k
                            break
                
                if cupo_key is not None:
                    cupo_categoria = int(dist[cupo_key])
                    ocupacion_categoria = await self.calcular_ocupacion_por_tipo(
                        db, zona_id, tipo_acceso, tipo_acceso_custom_id, inicio, fin
                    )
                    disponible_categoria = max(0, cupo_categoria - ocupacion_categoria)
                    
                    if cantidad > disponible_categoria:
                        exceso = cantidad - disponible_categoria
                        # Calcular cuántos puestos generales hay libres en la misma zona
                        ocupacion_zona = await self.calcular_ocupacion_proyectada(db, zona_id, inicio, fin)
                        cupo_zona_total = asig_zona.cupo_asignado
                        libre_zona_general = max(0, cupo_zona_total - ocupacion_zona)
                        puede_tomar_general = max(0, libre_zona_general - disponible_categoria)
                        
                        sug_nivel1 = []
                        if puede_tomar_general > 0:
                            tomar = min(exceso, puede_tomar_general)
                            sug_nivel1.append({
                                "accion": "tomar_general_zona",
                                "mensaje": f"Tomar {tomar} puestos del cupo general disponible en esta zona",
                                "cantidad_sugerida": min(disponible_categoria + tomar, cantidad),
                                "zona_id": str(zona_id),
                                "zona_nombre": asig_zona.zona_nombre
                            })
                        
                        # Buscar puestos del mismo tipo en otras zonas
                        for otra_asig in asigs:
                            if str(otra_asig.zona_id) == str(zona_id):
                                continue
                            otra_dist = otra_asig.distribucion_cupos or {}
                            for k, v in otra_dist.items():
                                k_lower = k.lower()
                                if k_lower == nombre_tipo.lower() or k == str(tipo_acceso_custom_id or ''):
                                    ocup_otra = await self.calcular_ocupacion_por_tipo(
                                        db, otra_asig.zona_id, tipo_acceso, tipo_acceso_custom_id, inicio, fin
                                    )
                                    libre_otra = max(0, int(v) - ocup_otra)
                                    if libre_otra > 0:
                                        sug_nivel1.append({
                                            "accion": "usar_otra_zona",
                                            "mensaje": f"Hay {libre_otra} puestos de {nombre_tipo} en '{otra_asig.zona_nombre}'",
                                            "cantidad_disponible": libre_otra,
                                            "zona_id": str(otra_asig.zona_id),
                                            "zona_nombre": otra_asig.zona_nombre
                                        })
                        
                        alertas.append({
                            "nivel": 1,
                            "tipo": "categoria_excedida",
                            "severidad": "warning",
                            "titulo": f"CUPO DE {nombre_tipo.upper()} EXCEDIDO",
                            "mensaje": f"Solo hay {disponible_categoria} puestos reservados para {nombre_tipo}. Solicitaste {cantidad} ({exceso} de exceso).",
                            "cupo_reservado": cupo_categoria,
                            "en_uso": ocupacion_categoria,
                            "disponible": disponible_categoria,
                            "sugerencias": sug_nivel1
                        })
        
        # ── NIVEL 2: Validación por Zona ──────────────────────────────────
        cupo_zona_disponible = None
        if zona_id:
            asig_zona = next((a for a in asigs if str(a.zona_id) == str(zona_id)), None)
            if asig_zona:
                ocupacion_zona = await self.calcular_ocupacion_proyectada(db, zona_id, inicio, fin)
                cupo_zona_total = asig_zona.cupo_asignado
                cupo_zona_disponible = max(0, cupo_zona_total - ocupacion_zona)
                
                if cantidad > cupo_zona_disponible:
                    exceso = cantidad - cupo_zona_disponible
                    
                    # Buscar disponibilidad en otras zonas
                    sug_nivel2 = []
                    for otra_asig in asigs:
                        if str(otra_asig.zona_id) == str(zona_id):
                            continue
                        ocup_otra = await self.calcular_ocupacion_proyectada(db, otra_asig.zona_id, inicio, fin)
                        libre_otra = max(0, otra_asig.cupo_asignado - ocup_otra)
                        if libre_otra > 0:
                            sug_nivel2.append({
                                "accion": "distribuir_otra_zona",
                                "mensaje": f"Distribuir {min(exceso, libre_otra)} pases en '{otra_asig.zona_nombre}' ({libre_otra} disponibles)",
                                "cantidad_disponible": libre_otra,
                                "zona_id": str(otra_asig.zona_id),
                                "zona_nombre": otra_asig.zona_nombre
                            })
                    
                    sug_nivel2.append({
                        "accion": "ajustar_cantidad",
                        "mensaje": f"Ajustar la cantidad a {cupo_zona_disponible} (máximo disponible en esta zona)",
                        "cantidad_sugerida": cupo_zona_disponible
                    })
                    
                    alertas.append({
                        "nivel": 2,
                        "tipo": "zona_excedida",
                        "severidad": "warning",
                        "titulo": "CAPACIDAD DE ZONA EXCEDIDA",
                        "mensaje": f"La zona tiene {cupo_zona_total} puestos asignados, {ocupacion_zona} comprometidos en este periodo. Disponibles: {cupo_zona_disponible}.",
                        "cupo_total": cupo_zona_total,
                        "en_uso": ocupacion_zona,
                        "disponible": cupo_zona_disponible,
                        "sugerencias": sug_nivel2
                    })
        
        # ── NIVEL 3: Validación Total Entidad (BLOQUEO DURO) ──────────────
        if cantidad > disponible_total_entidad:
            puede_crear = False
            alertas.append({
                "nivel": 3,
                "tipo": "entidad_excedida",
                "severidad": "error",
                "titulo": "CAPACIDAD TOTAL DE LA ENTIDAD SUPERADA",
                "mensaje": f"La entidad tiene {capacidad_total_entidad} puestos totales, {ocupacion_total_entidad} comprometidos. Máximo posible: {disponible_total_entidad}.",
                "cupo_total": capacidad_total_entidad,
                "en_uso": ocupacion_total_entidad,
                "disponible": disponible_total_entidad,
                "sugerencias": [{
                    "accion": "ajustar_cantidad",
                    "mensaje": f"Reducir a {disponible_total_entidad} (máximo de la entidad)",
                    "cantidad_sugerida": disponible_total_entidad
                }]
            })
        
        return {
            "puede_crear": puede_crear,
            "alertas": alertas,
            "resumen": {
                "cantidad_solicitada": cantidad,
                "cupo_total_entidad": capacidad_total_entidad,
                "disponible_total_entidad": disponible_total_entidad,
                "cupo_zona_disponible": cupo_zona_disponible,
                "cupo_categoria_disponible": disponible_categoria,
                "nombre_tipo_acceso": nombre_tipo,
            }
        }

    async def _generar_pases_simples(self, db: AsyncSession, lote: LotePaseMasivo, creado_por_id: uuid.UUID, extras: dict):
        """Genera N pases simples para el lote."""
        expira_at = datetime.combine(lote.fecha_fin, datetime.max.time()).replace(tzinfo=timezone.utc) + timedelta(hours=24)
        
        plan = extras.get('plan_distribucion')
        cursor_plan = 0
        puestos_en_zona_actual = 0
        
        for i in range(1, lote.cantidad_pases + 1):
            # Lógica de Distribución Inteligente
            zona_final_id = extras.get('zona_id') or lote.zona_estacionamiento_id
            if plan and cursor_plan < len(plan):
                zona_final_id = plan[cursor_plan]["zona_id"]
                puestos_en_zona_actual += 1
                if puestos_en_zona_actual >= plan[cursor_plan]["sugerencia"]:
                    cursor_plan += 1
                    puestos_en_zona_actual = 0

            serial_qr = f"{lote.codigo_serial}{str(i).zfill(4)}"
            token = crear_token_evento(serial_qr, expira_at)
            
            nuevo_qr = CodigoQR(
                token=token,
                tipo=QRTipo.evento_simple,
                lote_id=lote.id,
                serial_legible=serial_qr,
                max_accesos=lote.max_accesos_por_pase,
                fecha_expiracion=expira_at,
                created_by=creado_por_id,
                activo=True,
                # v2.0 campos
                tipo_acceso=extras.get('tipo_acceso', 'general'),
                tipo_acceso_custom_id=extras.get('tipo_acceso_custom_id'),
                zona_asignada_id=zona_final_id,
                puesto_asignado_id=extras.get('puesto_id'),
                multi_vehiculo=lote.multi_vehiculo,
                max_vehiculos=lote.max_vehiculos
            )
            db.add(nuevo_qr)

    async def _generar_pases_portal(self, db: AsyncSession, lote: LotePaseMasivo, creado_por_id: uuid.UUID, extras: dict):
        """Genera pre-usuarios para que se registren en el portal."""
        from app.core.security import hashear_password
        expira_at = datetime.combine(lote.fecha_fin, datetime.max.time()).replace(tzinfo=timezone.utc) + timedelta(hours=24)
        
        plan = extras.get('plan_distribucion')
        cursor_plan = 0
        puestos_en_zona_actual = 0

        for i in range(1, lote.cantidad_pases + 1):
            # Lógica de Distribución Inteligente
            zona_final_id = extras.get('zona_id') or lote.zona_estacionamiento_id
            if plan and cursor_plan < len(plan):
                zona_final_id = plan[cursor_plan]["zona_id"]
                puestos_en_zona_actual += 1
                if puestos_en_zona_actual >= plan[cursor_plan]["sugerencia"]:
                    cursor_plan += 1
                    puestos_en_zona_actual = 0

            serial_qr = f"{lote.codigo_serial}{str(i).zfill(4)}"
            token = crear_token_evento(serial_qr, expira_at)
            
            nuevo_qr = CodigoQR(
                token=token,
                tipo=QRTipo.evento_portal,
                lote_id=lote.id,
                serial_legible=serial_qr,
                max_accesos=lote.max_accesos_por_pase,
                fecha_expiracion=expira_at,
                created_by=creado_por_id,
                activo=True,
                tipo_acceso=extras.get('tipo_acceso', 'general'),
                tipo_acceso_custom_id=extras.get('tipo_acceso_custom_id'),
                zona_asignada_id=zona_final_id,
                puesto_asignado_id=extras.get('puesto_id'),
                multi_vehiculo=lote.multi_vehiculo,
                max_vehiculos=lote.max_vehiculos
            )
            db.add(nuevo_qr)
            await db.flush()

    async def procesar_excel_identificado(self, db: AsyncSession, lote: LotePaseMasivo, contenido_excel: bytes, creado_por_id: uuid.UUID):
        """Parsea Excel y crea pases identificados (Versión Backend Directa)."""
        import pandas as pd
        import io
        
        df = pd.read_excel(io.BytesIO(contenido_excel))
        df = df.where(pd.notnull(df), None)
        filas = df.values.tolist()
        await self.procesar_json_identificado(db, lote, filas, creado_por_id)

    async def procesar_json_identificado(self, db: AsyncSession, lote: LotePaseMasivo, filas: List[list], creado_por_id: uuid.UUID, extras: dict = None):
        """Parsea arreglo JSON (proveniente de Excel prevalidado) y crea pases identificados."""
        from app.models.codigo_qr import CodigoQR
        from app.models.vehiculo_pase import VehiculoPase
        from app.models.enums import QRTipo
        
        # 0. LIMPIEZA TÁCTICA: Si el lote ya tenía pases (ej. era tipo portal), los borramos 
        # para que la importación de Excel sea la fuente de verdad absoluta.
        await db.execute(delete(CodigoQR).where(CodigoQR.lote_id == lote.id))
        await db.flush()

        extras = extras or {}
        plan = extras.get('plan_distribucion')
        cursor_plan = 0
        puestos_en_zona_actual = 0

        # Límite de expiración
        expira_at = datetime.combine(lote.fecha_fin, datetime.max.time()).replace(tzinfo=timezone.utc) + timedelta(hours=24)
        count = 0
        
        for row in filas:
            if not row or not any(row): continue
            
            # Lógica de Distribución Inteligente
            zona_final_id = extras.get('zona_id') or lote.zona_estacionamiento_id
            if plan and cursor_plan < len(plan):
                zona_final_id = plan[cursor_plan]["zona_id"]
                puestos_en_zona_actual += 1
                if puestos_en_zona_actual >= plan[cursor_plan]["sugerencia"]:
                    cursor_plan += 1
                    puestos_en_zona_actual = 0

            # Nuevo Formato Excel (20 Col): [NOMBRE, CEDULA, EMAIL, TELEFONO, 
            # V1_PLACA, V1_MARCA, V1_MODELO, V1_COLOR, ...]
            row_data = (list(row) + [None]*20)[:20]
            nombre, cedula, email, telefono = row_data[0:4]
            v1_data = row_data[4:8]
            v2_data = row_data[8:12]
            v3_data = row_data[12:16]
            v4_data = row_data[16:20]
            
            serial_qr = f"{lote.codigo_serial}{str(count + 1).zfill(4)}"
            token = crear_token_evento(serial_qr, expira_at)
            
            # Verificar integridad de datos
            datos_estan_completos = bool(nombre and cedula and v1_data[0] and v1_data[1] and v1_data[2] and v1_data[3])
            
            nuevo_qr = CodigoQR(
                token=token,
                tipo=QRTipo.evento_identificado,
                lote_id=lote.id,
                serial_legible=serial_qr,
                max_accesos=lote.max_accesos_por_pase,
                fecha_expiracion=expira_at,
                created_by=creado_por_id,
                activo=True,
                nombre_portador=str(nombre).upper() if nombre else "INVITADO",
                cedula_portador=str(cedula).upper() if cedula else None,
                email_portador=str(email).lower() if email else None,
                telefono_portador=str(telefono) if telefono else None,
                # Vehículo 1
                vehiculo_placa=str(v1_data[0]).upper() if v1_data[0] else None,
                vehiculo_marca=str(v1_data[1]).upper() if v1_data[1] else None,
                vehiculo_modelo=str(v1_data[2]).upper() if v1_data[2] else None,
                vehiculo_color=str(v1_data[3]).upper() if v1_data[3] else None,
                tipo_acceso=lote.tipo_acceso,
                tipo_acceso_custom_id=lote.tipo_acceso_custom_id if hasattr(lote, 'tipo_acceso_custom_id') else None,
                zona_asignada_id=zona_final_id,
                multi_vehiculo=bool(v2_data[0] or v3_data[0] or v4_data[0]),
                max_vehiculos=lote.max_vehiculos,
                datos_completos=datos_estan_completos
            )
            db.add(nuevo_qr)
            await db.flush()
            
            # Vehículos adicionales (V2, V3, V4)
            for v_data in [v2_data, v3_data, v4_data]:
                if v_data[0]: # Si hay placa
                    v_extra = VehiculoPase(
                        qr_id=nuevo_qr.id,
                        placa=str(v_data[0]).upper(),
                        marca=str(v_data[1]).upper() if v_data[1] else None,
                        modelo=str(v_data[2]).upper() if v_data[2] else None,
                        color=str(v_data[3]).upper() if v_data[3] else None,
                        zona_asignada_id=zona_final_id
                    )
                    db.add(v_extra)
            
            count += 1
            
        lote.cantidad_pases = count
        # Nota: el commit se hace en crear_lote o en el endpoint

    def generar_qr_image(self, data: str, titulo: str = "", subtitulo: str = "", serial: str = "") -> io.BytesIO:
        """Genera una imagen QR de resolución balanceada (aprox 800x800) con texto descriptivo."""
        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=20, 
            border=4,
        )
        qr.add_data(data)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white").convert('RGB')
        
        if titulo or subtitulo or serial:
            from PIL import Image, ImageDraw, ImageFont
            qr_width, qr_height = img.size
            
            text_area_height = 80
            new_width = qr_width
            new_height = qr_height + text_area_height
            
            new_img = Image.new('RGB', (new_width, new_height), 'white')
            new_img.paste(img, (0, 0))
            
            draw = ImageDraw.Draw(new_img)
            
            import os
            font_path = os.path.join(os.path.dirname(__file__), "Roboto-Bold.ttf")
            font = None
            try:
                font = ImageFont.truetype(font_path, 36)
            except IOError:
                font = ImageFont.load_default()
                
            titulo_cortado = f"{titulo[:30]}..." if len(titulo) > 30 else titulo
            text = f"{titulo_cortado}  |  SERIAL: {serial}"
            
            if hasattr(font, 'getbbox'):
                bbox = draw.textbbox((0,0), text, font=font)
                text_width = bbox[2] - bbox[0]
                text_height = bbox[3] - bbox[1]
            else:
                try:
                    text_width, text_height = font.getsize(text)
                except Exception:
                    text_width, text_height = (len(text)*20, 30)
                
            text_x = max((new_width - text_width) // 2, 10)
            text_y = qr_height + (text_area_height - text_height) // 2 - 10
            
            draw.text((text_x, text_y), text, fill="black", font=font)
            
            img_byte_arr = io.BytesIO()
            new_img.save(img_byte_arr, format='PNG')
        else:
            img_byte_arr = io.BytesIO()
            img.save(img_byte_arr, format='PNG')
            
        img_byte_arr.seek(0)
        return img_byte_arr

    async def generar_zip_lote(self, db: AsyncSession, lote_id: uuid.UUID) -> str:
        """
        Genera un ZIP con todos los QRs del lote y lo sube a Supabase.
        Retorna la URL del archivo.
        """
        lote = await db.get(LotePaseMasivo, lote_id)
        if not lote: return None
        
        query = select(CodigoQR).where(CodigoQR.lote_id == lote.id)
        res = await db.execute(query)
        qrs = res.scalars().all()
        
        # --- AUTO-CURACIÓN TÁCTICA v3.2 ---
        # Si el lote existe pero no tiene QRs, y es de tipo generable (portal/simple), los creamos ahora.
        if not qrs:
            from app.models.enums import PasseTipo
            tipo_val = lote.tipo_pase.value if hasattr(lote.tipo_pase, 'value') else str(lote.tipo_pase)
            
            if tipo_val in [PasseTipo.portal.value, PasseTipo.simple.value]:
                # Re-generamos los pases mendiante los servicios correspondientes
                # El diccionario 'datos' requerido para extras se construye con lo básico del lote
                datos_reparacion = {
                    'zona_id': lote.zona_estacionamiento_id,
                    'tipo_acceso_custom_id': lote.tipo_acceso_custom_id,
                    'tipo_acceso': lote.tipo_acceso
                }
                
                if tipo_val == PasseTipo.portal.value:
                    await self._generar_pases_portal(db, lote, lote.creado_por, datos_reparacion)
                else:
                    await self._generar_pases_simples(db, lote, lote.creado_por, datos_reparacion)
                
                await db.commit()
                # Re-consultamos los QRs ahora generados
                res = await db.execute(query)
                qrs = res.scalars().all()
        # ----------------------------------

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'a', zipfile.ZIP_DEFLATED, False) as zip_file:
            from app.models.enums import QRTipo
            for qr in qrs:
                filename = f"{qr.serial_legible}.png"
                
                tipo_str = "QR GENÉRICO"
                if qr.tipo == QRTipo.evento_identificado:
                    tipo_str = "IDENTIFICADO"
                elif qr.tipo == QRTipo.evento_portal:
                    tipo_str = "AUTO-REGISTRO"
                    
                img_data = self.generar_qr_image(qr.token, lote.nombre_evento, tipo_str, qr.serial_legible)
                zip_file.writestr(filename, img_data.getvalue())
        
        zip_buffer.seek(0)
        
        # Subir a Supabase Storage
        if self.supabase:
            import re
            nombre_limpio = re.sub(r'[^a-zA-Z0-9_\-]', '_', lote.nombre_evento).strip('_')
            file_path = f"pases/{nombre_limpio}_{lote.codigo_serial}.zip"
            # Limpiar si ya existe
            try:
                self.supabase.storage.from_("bagfm-pases").remove([file_path])
            except: pass
            
            res_storage = self.supabase.storage.from_("bagfm-pases").upload(
                file_path, 
                zip_buffer.getvalue(),
                {"content-type": "application/zip"}
            )
            
            public_url = self.supabase.storage.from_("bagfm-pases").get_public_url(file_path)
            lote.zip_url = public_url
            lote.zip_generado = True
            lote.zip_listo_at = datetime.now(timezone.utc)
            await db.commit()
            return public_url
            
        return "storage_not_configured"

    async def generar_pdf_masivo(self, db: AsyncSession, lote_id: uuid.UUID) -> str:
        """
        Genera el PDF masivo del lote y lo sube a Supabase.
        Retorna la URL.
        """
        from app.services.pdf_service import pdf_service
        lote = await db.get(LotePaseMasivo, lote_id)
        if not lote: return None
        
        pdf_buffer = await pdf_service.generar_pdf_lote(db, lote_id)
        
        # Subir a Supabase
        if self.supabase:
            import re
            nombre_limpio = re.sub(r'[^a-zA-Z0-9_\-]', '_', lote.nombre_evento).strip('_')
            file_path = f"pases_pdf/{nombre_limpio}_{lote.codigo_serial}.pdf"
            
            try:
                self.supabase.storage.from_("bagfm-pases").remove([file_path])
            except: pass
            
            self.supabase.storage.from_("bagfm-pases").upload(
                file_path, 
                pdf_buffer.getvalue(),
                {"content-type": "application/pdf"}
            )
            
            public_url = self.supabase.storage.from_("bagfm-pases").get_public_url(file_path)
            lote.pdf_url = public_url # Asegurarse de que el modelo tiene este campo
            await db.commit()
            return public_url
            
        return "storage_not_configured"

    async def generar_excel_lote(self, db: AsyncSession, lote_id: uuid.UUID, base_url: str = "https://bagfm.app") -> io.BytesIO:
        """
        Genera un archivo Excel simplificado con hipervínculos tácticos directos al portal.
        """
        import pandas as pd
        from openpyxl.utils import get_column_letter
        
        lote = await db.get(LotePaseMasivo, lote_id)
        if not lote: return None
        
        query = select(CodigoQR).where(CodigoQR.lote_id == lote.id).order_by(CodigoQR.serial_legible.asc())
        res = await db.execute(query)
        qrs = res.scalars().all()
        
        data = []
        for qr in qrs:
            portal_url = f"{base_url}/portal/pase/{qr.token}"
            nombre = qr.nombre_portador or "INVITADO PENDIENTE"
            
            # Limpiar teléfono para link de WA (solo dígitos)
            tel_sucio = qr.telefono_portador or ""
            tel_limpio = "".join(filter(str.isdigit, tel_sucio))
            
            # Construir mensaje de WA
            msg = f"Estimado {nombre}, le enviamos su pase de acceso táctico para la Base Aérea Gral. Francisco de Miranda. Por favor complete sus datos aquí: {portal_url}"
            import urllib.parse
            wa_link = f"https://wa.me/{tel_limpio}?text={urllib.parse.quote(msg)}" if tel_limpio else ""

            data.append({
                "NOMBRE": nombre,
                "CEDULA": qr.cedula_portador or "POR REGISTRAR",
                "TELEFONO": tel_sucio or "S/T",
                "LINK PORTAL": portal_url,
                "COMPARTIR WHATSAPP": wa_link
            })
            
        df = pd.DataFrame(data)
        excel_buffer = io.BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='PASES_PORTAL')
            workbook = writer.book
            worksheet = writer.sheets['PASES_PORTAL']
            
            # Columnas con hipervínculos (D=4, E=5)
            col_links = [4, 5]
            textos = {4: "ACCEDER AL PORTAL", 5: "ENVIAR POR WHATSAPP"}
            
            from openpyxl.styles import Font
            link_font = Font(color="0000FF", underline="single")

            for row in range(2, len(data) + 2):
                for col_idx in col_links:
                    cell = worksheet.cell(row=row, column=col_idx)
                    link_url = cell.value
                    if link_url:
                        cell.hyperlink = link_url
                        cell.style = "Hyperlink"
                        cell.value = textos[col_idx]
                        cell.font = link_font

            # Auto-ajustar ancho de columnas
            for i, col in enumerate(df.columns):
                column_len = max(df[col].astype(str).str.len().max(), len(col)) + 5
                worksheet.column_dimensions[get_column_letter(i+1)].width = min(column_len, 60)
            
        excel_buffer.seek(0)
        return excel_buffer

    async def actualizar_pase(self, db: AsyncSession, pase_id: uuid.UUID, datos: dict) -> CodigoQR:
        """Actualiza datos de un pase individual (portador, vehículo, etc)."""
        from app.models.vehiculo_pase import VehiculoPase
        pase = await db.get(CodigoQR, pase_id)
        if not pase:
            return None
        
        # Extraer vehículos adicionales si vienen para reemplazo total (Admin mode)
        vehiculos_extra = datos.pop('vehiculos_adicionales', None)
        
        for key, value in datos.items():
            if hasattr(pase, key) and value is not None:
                setattr(pase, key, value)
        
        if vehiculos_extra is not None:
            # 1. Eliminar actuales
            await db.execute(delete(VehiculoPase).where(VehiculoPase.qr_id == pase.id))
            # 2. Insertar nuevos sanitizados
            for v in vehiculos_extra:
                if v.get('placa'):
                    nuevo_v = VehiculoPase(
                        qr_id=pase.id,
                        placa=self._sanitizar(v.get('placa'), 20, upper=True),
                        marca=self._sanitizar(v.get('marca'), 50, upper=True),
                        modelo=self._sanitizar(v.get('modelo'), 50, upper=True),
                        color=self._sanitizar(v.get('color'), 50, upper=True),
                        zona_asignada_id=pase.zona_asignada_id
                    )
                    db.add(nuevo_v)
        
        await db.commit()
        await db.refresh(pase)
        return pase

    async def eliminar_lote(self, db: AsyncSession, lote_id: uuid.UUID) -> bool:
        """
        Elimina un lote y toda su infraestructura asociada.
        - Borra QRs (cascada).
        - Borra Usuarios temporales (portal).
        - Borra archivos de Supabase (ZIP, PDF).
        - Borra el Lote.
        """
        lote = await db.get(LotePaseMasivo, lote_id)
        if not lote:
            return False
            
        # 1. Obtener IDs de usuarios temporales para este lote antes de borrar los QRs
        query_qrs = select(CodigoQR.usuario_id).where(CodigoQR.lote_id == lote_id)
        res_qrs = await db.execute(query_qrs)
        user_ids = [uid for uid in res_qrs.scalars().all() if uid]
        
        # 2. Eliminar archivos de Supabase
        if self.supabase:
            import re
            nombre_limpio = re.sub(r'[^a-zA-Z0-9_\-]', '_', lote.nombre_evento).strip('_')
            
            # Rutas posibles
            paths = [
                f"pases/{nombre_limpio}_{lote.codigo_serial}.zip",
                f"pases_pdf/{nombre_limpio}_{lote.codigo_serial}.pdf"
            ]
            
            try:
                self.supabase.storage.from_("bagfm-pases").remove(paths)
            except Exception as e:
                print(f"ALERTA STORAGE: No se pudieron borrar algunos archivos: {e}")
        
        # 3. Eliminar el lote (CodigoQR caen por CASCADE debido al modelo)
        await db.delete(lote)
        
        # Sincronizamos para que los QRs desaparezcan y se liberen las FKs de los Usuarios
        await db.flush()
        
        # 4. Eliminar Usuarios temporales asociados (Legacy Cleanup)
        # Solo borramos si el usuario NO tiene más registros en CodigoQR (usuarios huérfanos de este lote)
        if user_ids:
            from app.models.usuario import Usuario
            from sqlalchemy import update
            
            # Marcamos como eliminados los usuarios que NO tienen pases en otros lotes
            usuarios_con_pases = select(CodigoQR.usuario_id).where(CodigoQR.usuario_id != None)
            
            query_soft_del = update(Usuario).where(
                Usuario.id.in_(user_ids),
                Usuario.rol == RolTipo.SOCIO,
                ~Usuario.id.in_(usuarios_con_pases)
            ).values(is_deleted=True, activo=False)
            
            await db.execute(query_soft_del)
            
        await db.commit()
        return True

    async def registrar_invitado_portal(self, db: AsyncSession, lote_id: uuid.UUID, datos: dict) -> dict:
        """
        Registra un invitado en un lote tipo PORTAL.
        Busca un pase 'vacío' (sin nombre) o crea uno nuevo si hay cupo.
        """
        lote = await db.get(LotePaseMasivo, lote_id)
        if not lote:
            raise ValueError("Lote no encontrado")

        # 1. Buscar si hay pases libres en este lote
        query = select(CodigoQR).where(
            CodigoQR.lote_id == lote_id,
            CodigoQR.nombre_portador == None,
            CodigoQR.activo == True
        ).limit(1)
        res = await db.execute(query)
        pase = res.scalar_one_or_none()

        if not pase:
            # Si no hay libres, verificar si podemos crear uno nuevo (cupo dinámico)
            # Por ahora, asumimos que el lote pre-generó todos sus QRs.
            raise ValueError("No hay cupos disponibles para este lote")

        # 2. Actualizar datos del portador con sanitización táctica
        pase.nombre_portador = self._sanitizar(datos.get('nombre'), 100, upper=True)
        pase.cedula_portador = self._sanitizar(datos.get('cedula'), 20, upper=True)
        pase.email_portador = self._sanitizar(datos.get('email'), 100, lower=True)
        pase.telefono_portador = self._sanitizar(datos.get('telefono'), 20)
        
        # 3. Datos del Vehículo (Sanitizados)
        pase.vehiculo_placa = self._sanitizar(datos.get('placa'), 15, upper=True)
        pase.vehiculo_marca = self._sanitizar(datos.get('marca'), 30, upper=True)
        pase.vehiculo_modelo = self._sanitizar(datos.get('modelo'), 30, upper=True)
        pase.vehiculo_color = self._sanitizar(datos.get('color'), 20, upper=True)
        
        pase.tipo_acceso = lote.tipo_acceso
        pase.tipo_acceso_custom_id = lote.tipo_acceso_custom_id
        pase.datos_completos = True
        
        await db.commit()
        await db.refresh(pase)
        
        # Generar QR base64 para mostrar al usuario inmediatamente
        qr_base64 = self.generar_qr_base64(pase.token)
        
        return {
            "id": pase.id,
            "serial": pase.serial_legible,
            "qr_base64": qr_base64
        }

    async def actualizar_pase_publico(self, db: AsyncSession, pase_id: uuid.UUID, datos: dict) -> dict:
        """Permite completar datos de un pase pre-identificado (Excel parcial)."""
        from app.models.vehiculo_pase import VehiculoPase
        pase = await db.get(CodigoQR, pase_id)
        if not pase:
            raise ValueError("Pase no encontrado")
            
        # Actualización portador
        if datos.get('nombre'): pase.nombre_portador = self._sanitizar(datos['nombre'], 100, upper=True)
        if datos.get('cedula'): pase.cedula_portador = self._sanitizar(datos['cedula'], 20, upper=True)
        if datos.get('email'): pase.email_portador = self._sanitizar(datos['email'], 100, lower=True)
        if datos.get('telefono'): pase.telefono_portador = self._sanitizar(datos['telefono'], 20)
        
        # Vehículo 1 (QR principal)
        if datos.get('placa'): pase.vehiculo_placa = self._sanitizar(datos['placa'], 15, upper=True)
        if datos.get('marca'): pase.vehiculo_marca = self._sanitizar(datos['marca'], 30, upper=True)
        if datos.get('modelo'): pase.vehiculo_modelo = self._sanitizar(datos['modelo'], 30, upper=True)
        if datos.get('color'): pase.vehiculo_color = self._sanitizar(datos['color'], 20, upper=True)
        
        # Vehículos Adicionales (Portal solo permite EDITAR los existentes del Excel)
        vehiculos_extra_data = datos.get('vehiculos_adicionales', [])
        if vehiculos_extra_data:
            query = select(VehiculoPase).where(VehiculoPase.qr_id == pase.id).order_by(VehiculoPase.created_at.asc())
            res = await db.execute(query)
            actuales = res.scalars().all()
            
            # Sincronizamos solo hasta el número de vehículos que ya existían
            for i, v_data in enumerate(vehiculos_extra_data):
                if i < len(actuales):
                    v_db = actuales[i]
                    if v_data.get('placa'): v_db.placa = self._sanitizar(v_data['placa'], 20, upper=True)
                    if v_data.get('marca'): v_db.marca = self._sanitizar(v_data['marca'], 50, upper=True)
                    if v_data.get('modelo'): v_db.modelo = self._sanitizar(v_data['modelo'], 50, upper=True)
                    if v_data.get('color'): v_db.color = self._sanitizar(v_data['color'], 50, upper=True)

        pase.datos_completos = True
        await db.commit()
        await db.refresh(pase)
        
        return {"status": "ok", "serial": pase.serial_legible}

    def _sanitizar(self, texto: any, max_len: int, upper: bool = False, lower: bool = False) -> str:
        """Limpia el texto de etiquetas HTML, scripts y limita longitud para evitar ataques XSS o desbordamientos."""
        if texto is None: return None
        s = str(texto).strip()
        
        # Eliminar etiquetas HTML de forma agresiva
        import re
        s = re.sub(r'<[^>]*?>', '', s)
        
        # Limitar longitud
        s = s[:max_len]
        
        if upper: s = s.upper()
        if lower: s = s.lower()
        return s

pase_service = PaseService()
