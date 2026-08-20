"""
Servicio del módulo Dormitorios.

Aquí vive la única regla que no puede quedarse en la capa HTTP: **el integrante no es
una tabla nueva de personas**. Cada alta busca primero en `usuarios` por cédula y, si
la persona ya está, le cuelga el perfil militar en vez de duplicarla. Ese es el motivo
de existir del módulo: que registrar a alguien porque duerme en la base lo ate al
vehículo que ya tenía a su nombre, sin que nadie cruce nada a mano.
"""
import io
import logging
from datetime import datetime, timezone
from typing import List, Optional
from uuid import UUID

from sqlalchemy import func, or_, select, update
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.excepciones import CapacidadExcedida, EntidadDuplicada, EntidadNoEncontrada
from app.core.security import crear_token_qr, hashear_password
from app.core.tokens import generar_token, hashear_token
from app.models.codigo_qr import CodigoQR
from app.models.dormitorio import Dormitorio, Habitacion
from app.models.enums import QRTipo, RolTipo
from app.models.perfil_militar import PerfilMilitar
from app.models.usuario import Usuario
from app.models.vehiculo import Vehiculo
from app.services.placa_lookup import _placa_sin_separadores, normalizar_placa
from app.schemas.dormitorio import (
    HabitacionPublica,
    IntegranteCrear,
    IntegranteEditar,
    IntegranteSalida,
    OcupantePublico,
    VehiculoResumen,
)

logger = logging.getLogger(__name__)


class DormitorioService:

    # ─── Consultas de apoyo ────────────────────────────────────────────────────

    async def obtener_dormitorio(self, db: AsyncSession, dormitorio_id: UUID) -> Dormitorio:
        dormitorio = await db.get(Dormitorio, dormitorio_id)
        if not dormitorio:
            raise EntidadNoEncontrada("Dormitorio no encontrado")
        return dormitorio

    async def obtener_habitacion(self, db: AsyncSession, habitacion_id: UUID) -> Habitacion:
        habitacion = await db.get(Habitacion, habitacion_id)
        if not habitacion:
            raise EntidadNoEncontrada("Habitación no encontrada")
        return habitacion

    async def obtener_perfil(self, db: AsyncSession, usuario_id: UUID) -> PerfilMilitar:
        res = await db.execute(
            select(PerfilMilitar).where(PerfilMilitar.usuario_id == usuario_id)
        )
        perfil = res.scalars().first()
        if not perfil:
            raise EntidadNoEncontrada("Este usuario no tiene perfil de integrante")
        return perfil

    # ─── Ficha del integrante ──────────────────────────────────────────────────

    async def construir_integrante(self, db: AsyncSession, perfil: PerfilMilitar) -> IntegranteSalida:
        """
        Une las tres fuentes que describen a un residente: su fila de `usuarios`, su
        perfil militar y los vehículos que constan realmente a su nombre.
        """
        usuario = perfil.usuario or await db.get(Usuario, perfil.usuario_id)

        res_veh = await db.execute(
            select(Vehiculo).where(
                Vehiculo.socio_id == perfil.usuario_id,
                Vehiculo.activo == True,  # noqa: E712
            )
        )
        vehiculos = [VehiculoResumen.model_validate(v) for v in res_veh.scalars().all()]

        # ¿Tiene ya carnet peatonal emitido? Sirve para no reimprimir sin querer.
        res_qr = await db.execute(
            select(func.count(CodigoQR.id)).where(
                CodigoQR.usuario_id == perfil.usuario_id,
                CodigoQR.tipo == QRTipo.permanente,
                CodigoQR.activo == True,  # noqa: E712
            )
        )
        tiene_qr = (res_qr.scalar() or 0) > 0

        habitacion = perfil.habitacion
        dormitorio = habitacion.dormitorio if habitacion else None

        return IntegranteSalida(
            usuario_id=perfil.usuario_id,
            cedula=usuario.cedula if usuario else "",
            nombre=usuario.nombre if usuario else "",
            apellido=usuario.apellido if usuario else "",
            telefono=usuario.telefono if usuario else None,
            email=usuario.email if usuario else None,
            grado=perfil.grado,
            unidad=perfil.unidad,
            jefe_nombre=perfil.jefe_nombre,
            jefe_telefono=perfil.jefe_telefono,
            tiene_vehiculo=perfil.tiene_vehiculo,
            vehiculos=vehiculos,
            habitacion_id=perfil.habitacion_id,
            habitacion_numero=habitacion.numero if habitacion else None,
            dormitorio_id=dormitorio.id if dormitorio else None,
            dormitorio_nombre=dormitorio.nombre if dormitorio else None,
            tiene_qr_peatonal=tiene_qr,
            activo=perfil.activo,
        )

    async def listar_integrantes_habitacion(
        self, db: AsyncSession, habitacion_id: UUID
    ) -> List[IntegranteSalida]:
        res = await db.execute(
            select(PerfilMilitar)
            .options(selectinload(PerfilMilitar.usuario))
            .where(PerfilMilitar.habitacion_id == habitacion_id, PerfilMilitar.activo == True)  # noqa: E712
        )
        return [await self.construir_integrante(db, p) for p in res.scalars().all()]

    async def habitaciones_paginadas(
        self,
        db: AsyncSession,
        dormitorio_id: UUID,
        pagina: int = 1,
        por_pagina: int = 10,
        buscar: str = "",
    ) -> tuple[List[Habitacion], int]:
        """
        Habitaciones del dormitorio, por páginas y con filtro por persona.

        El filtro busca en la GENTE, no en el número de habitación: la pregunta real
        del Comandante es "¿dónde duerme Rojas?", no "¿existe la 104?". Devuelve las
        habitaciones que alojan a alguien que coincide por nombre, apellido, cédula o
        grado, y la paginación se aplica después del filtro para que el número de
        páginas se corresponda con lo que se está viendo.
        """
        base = select(Habitacion).where(
            Habitacion.dormitorio_id == dormitorio_id,
            Habitacion.activo == True,  # noqa: E712
        )

        termino = (buscar or "").strip()
        if termino:
            patron = f"%{termino.lower()}%"
            coincidentes = (
                select(PerfilMilitar.habitacion_id)
                .join(Usuario, PerfilMilitar.usuario_id == Usuario.id)
                .where(
                    PerfilMilitar.activo == True,  # noqa: E712
                    or_(
                        func.lower(Usuario.nombre).like(patron),
                        func.lower(Usuario.apellido).like(patron),
                        func.lower(Usuario.cedula).like(patron),
                        func.lower(PerfilMilitar.grado).like(patron),
                    ),
                )
            )
            base = base.where(Habitacion.id.in_(coincidentes))

        total = (await db.execute(
            select(func.count()).select_from(base.subquery())
        )).scalar() or 0

        pagina = max(1, pagina)
        por_pagina = max(1, min(por_pagina, 100))

        res = await db.execute(
            base.order_by(
                func.coalesce(func.length(Habitacion.piso), 0),
                Habitacion.piso.asc().nulls_first(),
                func.length(Habitacion.numero),
                Habitacion.numero,
            )
            .offset((pagina - 1) * por_pagina)
            .limit(por_pagina)
        )
        return list(res.scalars().all()), total

    # ─── Alta y edición de integrantes ─────────────────────────────────────────

    async def crear_integrante(
        self, db: AsyncSession, datos: IntegranteCrear, creador_id: Optional[UUID] = None
    ) -> PerfilMilitar:
        cedula = datos.cedula.strip().upper()

        res = await db.execute(select(Usuario).where(Usuario.cedula == cedula))
        usuario = res.scalars().first()

        if usuario:
            # La persona ya está en el sistema (socio de un club, guardia, quien sea).
            # Se completa, no se duplica: ese es todo el objetivo del módulo.
            #
            # El ROL no se toca. Si la cédula resulta ser la del Comandante porque
            # alguien lo apuntó en un dormitorio, degradarlo a SOCIO lo dejaría fuera
            # de su propio panel.
            usuario.nombre = datos.nombre.strip()
            usuario.apellido = datos.apellido.strip()
            if datos.telefono:
                usuario.telefono = datos.telefono.strip()
            if datos.email:
                usuario.email = datos.email.strip()
            usuario.is_deleted = False
            usuario.activo = True
        else:
            if datos.email:
                res_email = await db.execute(
                    select(Usuario).where(func.lower(Usuario.email) == datos.email.strip().lower())
                )
                if res_email.scalars().first():
                    raise EntidadDuplicada(f"El correo {datos.email} ya está registrado por otro usuario")

            usuario = Usuario(
                cedula=cedula,
                nombre=datos.nombre.strip(),
                apellido=datos.apellido.strip(),
                email=datos.email.strip() if datos.email else None,
                telefono=datos.telefono.strip() if datos.telefono else None,
                # SOCIO es el rol "persona sin panel" que el sistema ya usa. No se crea
                # un rol nuevo: `rol_tipo` es un ENUM nativo de Postgres y alterarlo ya
                # costó caro una vez (ver fix_enum.py en la raíz del repo).
                rol=RolTipo.SOCIO,
                password_hash=hashear_password(cedula),
                debe_cambiar_password=True,
            )
            db.add(usuario)
            await db.flush()

        # Un integrante no lleva membresía: no es socio de ningún club, duerme aquí.

        res_perfil = await db.execute(
            select(PerfilMilitar).where(PerfilMilitar.usuario_id == usuario.id)
        )
        perfil = res_perfil.scalars().first()

        if perfil is None:
            perfil = PerfilMilitar(usuario_id=usuario.id)
            db.add(perfil)

        perfil.grado = datos.grado
        perfil.unidad = datos.unidad
        perfil.jefe_nombre = datos.jefe_nombre
        perfil.jefe_telefono = datos.jefe_telefono
        perfil.activo = True

        # La declaración solo puede ser "no tiene vehículo" si de verdad no consta
        # ninguno a su nombre.
        #
        # `tiene_vehiculo` está para hacer visible una discrepancia REAL: alguien que
        # niega tener vehículo y sin embargo aparece con placas. Guardar un "no" cuando
        # el propio sistema ya tiene la placa registrada no descubre nada — fabrica la
        # contradicción, y la pantalla acaba enseñando la placa junto al reproche de
        # haberla negado. La discrepancia tiene que venir de los datos, no de este
        # formulario.
        res_veh_suyos = await db.execute(
            select(func.count(Vehiculo.id)).where(
                Vehiculo.socio_id == usuario.id,
                Vehiculo.activo == True,  # noqa: E712
            )
        )
        tiene_placas_reales = (res_veh_suyos.scalar() or 0) > 0
        perfil.tiene_vehiculo = datos.tiene_vehiculo or tiene_placas_reales

        if datos.habitacion_id:
            habitacion = await self.obtener_habitacion(db, datos.habitacion_id)
            await self._verificar_camas(db, habitacion, excluir_usuario_id=usuario.id)
            perfil.habitacion_id = habitacion.id
            perfil.fecha_ingreso_dormitorio = datetime.now(timezone.utc)

        await db.commit()
        await db.refresh(perfil)
        return perfil

    async def editar_integrante(
        self, db: AsyncSession, usuario_id: UUID, datos: IntegranteEditar
    ) -> PerfilMilitar:
        perfil = await self.obtener_perfil(db, usuario_id)
        usuario = await db.get(Usuario, usuario_id)

        cambios = datos.model_dump(exclude_unset=True)

        for campo in ("nombre", "apellido", "telefono", "email"):
            if campo in cambios and usuario:
                setattr(usuario, campo, cambios[campo])

        for campo in ("grado", "unidad", "jefe_nombre", "jefe_telefono", "tiene_vehiculo", "activo"):
            if campo in cambios:
                setattr(perfil, campo, cambios[campo])

        # Dar de baja al integrante libera la cama: dejarlo ocupando una plaza que ya no
        # usa falsearía la capacidad del dormitorio entero.
        if cambios.get("activo") is False:
            perfil.habitacion_id = None

        await db.commit()
        await db.refresh(perfil)
        return perfil

    # ─── Camas ─────────────────────────────────────────────────────────────────

    async def _verificar_camas(
        self, db: AsyncSession, habitacion: Habitacion, excluir_usuario_id: Optional[UUID] = None
    ) -> None:
        """
        La ocupación se cuenta contra la base y no sobre la relación ya cargada: dos
        administradores asignando a la vez verían cada uno su propia copia en memoria.
        """
        condiciones = [
            PerfilMilitar.habitacion_id == habitacion.id,
            PerfilMilitar.activo == True,  # noqa: E712
        ]
        if excluir_usuario_id:
            condiciones.append(PerfilMilitar.usuario_id != excluir_usuario_id)

        res = await db.execute(select(func.count(PerfilMilitar.id)).where(*condiciones))
        ocupacion = res.scalar() or 0

        if ocupacion >= habitacion.camas:
            raise CapacidadExcedida(
                f"La habitación {habitacion.numero} tiene {habitacion.camas} cama(s) y "
                f"{ocupacion} ocupada(s). Cambia el número de camas o libera una plaza "
                f"antes de asignar."
            )

    async def asignar_habitacion(
        self, db: AsyncSession, habitacion_id: UUID, usuario_id: UUID
    ) -> PerfilMilitar:
        habitacion = await self.obtener_habitacion(db, habitacion_id)
        perfil = await self.obtener_perfil(db, usuario_id)

        if perfil.habitacion_id == habitacion.id:
            return perfil

        await self._verificar_camas(db, habitacion, excluir_usuario_id=usuario_id)

        perfil.habitacion_id = habitacion.id
        perfil.fecha_ingreso_dormitorio = datetime.now(timezone.utc)
        await db.commit()
        await db.refresh(perfil)
        return perfil

    async def desasignar_habitacion(self, db: AsyncSession, usuario_id: UUID) -> PerfilMilitar:
        perfil = await self.obtener_perfil(db, usuario_id)
        perfil.habitacion_id = None
        await db.commit()
        await db.refresh(perfil)
        return perfil

    # ─── QR de la puerta ───────────────────────────────────────────────────────

    async def obtener_token_habitacion(
        self, db: AsyncSession, habitacion_id: UUID, regenerar: bool = False
    ) -> tuple[Habitacion, str]:
        """
        Devuelve el QR de la puerta. Lo crea la primera vez y lo conserva después.

        Consultarlo NO lo cambia. Eso es deliberado y corrige el comportamiento
        anterior: el QR está impreso y pegado en una puerta, así que bastaba con
        abrir el panel a mirarlo para dejar el adhesivo muerto sin que nadie se
        enterase hasta que alguien lo escaneara.

        `regenerar=True` sí emite uno nuevo, y ahí sí hay que imprimirlo y cambiarlo
        físicamente: el anterior deja de servir en la petición siguiente. Es la vía
        para cortar el acceso cuando se sabe que alguien fotografió la puerta.
        """
        habitacion = await self.obtener_habitacion(db, habitacion_id)

        if habitacion.token and not regenerar:
            return habitacion, habitacion.token

        token = generar_token()
        habitacion.token = token
        habitacion.token_hash = hashear_token(token)
        habitacion.token_pista = token[:6]
        habitacion.token_generado_at = datetime.now(timezone.utc)
        await db.commit()
        await db.refresh(habitacion)
        return habitacion, token

    async def revocar_token_habitacion(self, db: AsyncSession, habitacion_id: UUID) -> Habitacion:
        habitacion = await self.obtener_habitacion(db, habitacion_id)
        habitacion.token = None
        habitacion.token_hash = None
        habitacion.token_pista = None
        habitacion.token_generado_at = None
        await db.commit()
        await db.refresh(habitacion)
        return habitacion

    async def ficha_publica(self, db: AsyncSession, token: str) -> HabitacionPublica:
        """
        Lo que ve quien escanea el QR de la puerta. Sin sesión.

        Un token inexistente, uno revocado y una habitación desactivada dan todos el
        mismo error: distinguirlos convertiría el endpoint en un oráculo para averiguar
        qué tokens existen.
        """
        res = await db.execute(
            select(Habitacion)
            .options(selectinload(Habitacion.dormitorio))
            .where(Habitacion.token_hash == hashear_token(token))
        )
        habitacion = res.scalars().first()

        if not habitacion or not habitacion.activo:
            raise EntidadNoEncontrada("Habitación no encontrada")
        if habitacion.dormitorio and not habitacion.dormitorio.activo:
            raise EntidadNoEncontrada("Habitación no encontrada")

        res_ocup = await db.execute(
            select(PerfilMilitar)
            .options(selectinload(PerfilMilitar.usuario))
            .where(PerfilMilitar.habitacion_id == habitacion.id, PerfilMilitar.activo == True)  # noqa: E712
        )
        perfiles = res_ocup.scalars().all()

        ocupantes = []
        for p in perfiles:
            u = p.usuario
            if not u:
                continue
            ocupantes.append(
                OcupantePublico(
                    grado=p.grado,
                    nombre=u.nombre,
                    apellido=u.apellido,
                    unidad=p.unidad,
                    telefono=u.telefono,
                    jefe_nombre=p.jefe_nombre,
                    jefe_telefono=p.jefe_telefono,
                )
            )

        return HabitacionPublica(
            dormitorio=habitacion.dormitorio.nombre if habitacion.dormitorio else "",
            habitacion=habitacion.numero,
            piso=habitacion.piso,
            camas=habitacion.camas,
            ocupacion=len(ocupantes),
            ocupantes=ocupantes,
        )

    # ─── QR peatonal ───────────────────────────────────────────────────────────

    async def obtener_qr_peatonal(
        self,
        db: AsyncSession,
        usuario_id: UUID,
        creador_id: Optional[UUID] = None,
        regenerar: bool = False,
    ) -> tuple[Usuario, PerfilMilitar, str]:
        """
        Carnet QR para quien entra a pie. Se emite una vez y se conserva.

        Va en `codigos_qr` como cualquier otro pase —mismo escáner, misma bitácora— con
        `vehiculo_id` en nulo porque no hay vehículo que registrar, y sin membresía: lo
        que lo valida es la rama de perfil militar de `acceso_service.validar_qr`.

        Consultarlo devuelve el mismo carnet que ya tiene la persona, de modo que
        reimprimirlo por deterioro no deja fuera al que lleva en el bolsillo. Solo
        `regenerar=True` emite uno nuevo, y entonces sí anula el anterior — que es lo
        que hace falta cuando alguien pierde la credencial.
        """
        perfil = await self.obtener_perfil(db, usuario_id)
        usuario = await db.get(Usuario, usuario_id)
        if not usuario:
            raise EntidadNoEncontrada("Usuario no encontrado")

        if not regenerar:
            res = await db.execute(
                select(CodigoQR)
                .where(
                    CodigoQR.usuario_id == usuario_id,
                    CodigoQR.tipo == QRTipo.permanente,
                    CodigoQR.activo == True,  # noqa: E712
                )
                .order_by(CodigoQR.created_at.desc())
            )
            vigente = res.scalars().first()
            if vigente:
                return usuario, perfil, vigente.token

        # Un carnet nuevo deja sin valor al anterior: si no, quien perdió la credencial
        # sigue teniendo una válida circulando por ahí.
        await db.execute(
            update(CodigoQR)
            .where(CodigoQR.usuario_id == usuario_id, CodigoQR.activo == True)  # noqa: E712
            .values(activo=False)
        )

        # `unico=True` porque aquí sí se dan las dos emisiones seguidas: emitir el
        # carnet y regenerarlo acto seguido cae en el mismo segundo, y sin esto el JWT
        # sería idéntico y chocaría contra el índice único de `codigos_qr.token`.
        token = crear_token_qr(str(usuario_id), unico=True)
        qr = CodigoQR(
            usuario_id=usuario_id,
            tipo=QRTipo.permanente,
            token=token,
            activo=True,
            nombre_portador=f"{usuario.nombre} {usuario.apellido}",
            cedula_portador=usuario.cedula,
            telefono_portador=usuario.telefono,
            created_by=creador_id,
        )
        db.add(qr)
        await db.commit()
        return usuario, perfil, token

    # ─── Búsquedas que alimentan el formulario de alta ─────────────────────────

    async def buscar_personas(self, db: AsyncSession, termino: str) -> List[dict]:
        """
        Personas de `usuarios` que encajan con lo tecleado, para autorrellenar el alta.

        Compara la cédula **sin la letra ni los separadores**: en la base conviven
        "V12345678", "12345678" y "V-12.345.678" para la misma persona, y si el
        formulario solo casara la cadena exacta se crearían fichas duplicadas justo en
        el módulo que existe para unificarlas.
        """
        termino = (termino or "").strip()
        if len(termino) < 2:
            return []

        patron = f"%{termino.lower()}%"
        solo_digitos = "".join(c for c in termino if c.isdigit())

        condiciones = [
            func.lower(Usuario.nombre).like(patron),
            func.lower(Usuario.apellido).like(patron),
            func.lower(Usuario.cedula).like(patron),
        ]
        if solo_digitos:
            # Quita letra, guiones y puntos de la columna antes de comparar.
            cedula_normalizada = func.replace(
                func.replace(
                    func.replace(func.lower(Usuario.cedula), "v", ""), "-", ""
                ),
                ".",
                "",
            )
            condiciones.append(cedula_normalizada.like(f"%{solo_digitos}%"))

        res = await db.execute(
            select(Usuario)
            .where(Usuario.is_deleted == False, or_(*condiciones))  # noqa: E712
            .order_by(Usuario.apellido, Usuario.nombre)
            .limit(15)
        )
        personas = res.scalars().all()
        if not personas:
            return []

        ids = [p.id for p in personas]
        res_perfiles = await db.execute(
            select(PerfilMilitar).where(PerfilMilitar.usuario_id.in_(ids))
        )
        perfiles = {p.usuario_id: p for p in res_perfiles.scalars().all()}

        # Los vehículos que ya constan a su nombre viajan con la sugerencia. Sin esto,
        # elegir a alguien que ya tiene placa registrada dejaba el formulario diciendo
        # que no tiene vehículo, y al guardar la ficha salía con la placa Y con el aviso
        # de que lo había negado — una contradicción fabricada por el propio formulario.
        res_veh = await db.execute(
            select(Vehiculo).where(
                Vehiculo.socio_id.in_(ids),
                Vehiculo.activo == True,  # noqa: E712
            )
        )
        vehiculos_por_persona: dict = {}
        for v in res_veh.scalars().all():
            vehiculos_por_persona.setdefault(v.socio_id, []).append(v)

        salida = []
        for p in personas:
            perfil = perfiles.get(p.id)
            sus_vehiculos = vehiculos_por_persona.get(p.id, [])
            salida.append({
                "usuario_id": p.id,
                "cedula": p.cedula,
                "nombre": p.nombre,
                "apellido": p.apellido,
                "telefono": p.telefono,
                "email": p.email,
                "rol": p.rol.value if p.rol else None,
                # Lo que ya sabemos de ella, para no obligar a reescribirlo.
                "grado": perfil.grado if perfil else None,
                "unidad": perfil.unidad if perfil else None,
                "jefe_nombre": perfil.jefe_nombre if perfil else None,
                "jefe_telefono": perfil.jefe_telefono if perfil else None,
                # Si ya duerme en otra habitación, el alta sería un traslado.
                "ya_es_integrante": perfil is not None and perfil.activo,
                "habitacion_id": perfil.habitacion_id if perfil else None,
                # Lo que ya tiene a su nombre, para que el formulario no le pregunte
                # por algo que el sistema ya sabe.
                "vehiculos": [
                    {
                        "id": v.id,
                        "placa": v.placa,
                        "marca": v.marca,
                        "modelo": v.modelo,
                        "color": v.color,
                    }
                    for v in sus_vehiculos
                ],
            })
        return salida

    async def buscar_jefes(self, db: AsyncSession, termino: str) -> List[dict]:
        """
        Candidatos a jefe directo entre las personas ya registradas.

        El jefe se guarda como texto y no como clave ajena —muchas veces no está en el
        sistema y exigirlo convertiría un alta en un alta en cadena—, pero cuando sí
        está, elegirlo de la lista trae su teléfono y evita tres grafías del mismo
        apellido.
        """
        termino = (termino or "").strip()
        if len(termino) < 2:
            return []

        patron = f"%{termino.lower()}%"
        res = await db.execute(
            select(Usuario)
            .where(
                Usuario.is_deleted == False,  # noqa: E712
                or_(
                    func.lower(Usuario.nombre).like(patron),
                    func.lower(Usuario.apellido).like(patron),
                ),
            )
            .order_by(Usuario.apellido, Usuario.nombre)
            .limit(15)
        )
        personas = res.scalars().all()

        ids = [p.id for p in personas]
        grados = {}
        if ids:
            res_perfiles = await db.execute(
                select(PerfilMilitar).where(PerfilMilitar.usuario_id.in_(ids))
            )
            grados = {p.usuario_id: p.grado for p in res_perfiles.scalars().all()}

        return [
            {
                "usuario_id": p.id,
                "nombre_completo": p.nombre_completo,
                "telefono": p.telefono,
                "grado": grados.get(p.id),
            }
            for p in personas
        ]

    async def buscar_vehiculos(
        self, db: AsyncSession, placa: str, usuario_id: Optional[UUID] = None
    ) -> List[dict]:
        """
        Vehículos cuya placa se parece a lo tecleado, con su dueño si lo tienen.

        Devuelve también los que YA tienen dueño, marcados como no asignables. Ocultarlos
        haría creer que la placa está libre y llevaría a registrarla otra vez, que es
        exactamente cómo aparecieron las fichas duplicadas que hubo que sanear.

        `usuario_id` distingue el caso que importa: un vehículo que ya está a nombre de
        LA MISMA persona que se está registrando no es una placa ajena, es la suya, y
        tiene que poder elegirla. Sin esta distinción el sistema le impedía declarar su
        propio vehículo y después le reprochaba no haberlo declarado.
        """
        placa = normalizar_placa(placa or "")
        if len(placa) < 2:
            return []

        # Se compara sobre la columna ya normalizada, no sobre el texto crudo: hay
        # placas cargadas por Excel como "AV-645" que nunca casarían con "AV645".
        res = await db.execute(
            select(Vehiculo)
            .options(selectinload(Vehiculo.socio))
            .where(
                _placa_sin_separadores(Vehiculo.placa).like(f"%{placa}%"),
                Vehiculo.activo == True,  # noqa: E712
            )
            .order_by(Vehiculo.placa)
            .limit(15)
        )

        salida = []
        for v in res.scalars().all():
            es_suyo = usuario_id is not None and v.socio_id == usuario_id
            salida.append({
                "id": v.id,
                "placa": v.placa,
                "marca": v.marca,
                "modelo": v.modelo,
                "color": v.color,
                "libre": v.socio_id is None,
                "es_suyo": es_suyo,
                # El nombre del dueño solo interesa cuando es OTRA persona: es lo que
                # explica por qué esa placa no se puede tomar desde aquí.
                "dueno": v.socio.nombre_completo if (v.socio and not es_suyo) else None,
            })
        return salida

    async def vincular_vehiculo(
        self,
        db: AsyncSession,
        usuario_id: UUID,
        vehiculo_id: Optional[UUID] = None,
        datos_nuevos: Optional[dict] = None,
    ) -> Vehiculo:
        """
        Ata una placa a un integrante, en la tabla madre `vehiculos`.

        Solo acepta un vehículo **sin dueño**. Reasignar uno que ya tiene titular desde
        aquí sería quitárselo a alguien sin dejar rastro, y ese cambio pertenece al
        módulo de Parque Automotor, que es donde se ve el historial completo.

        Si la placa no existe se crea la fila aquí mismo, no en una tabla aparte: así la
        cámara de la alcabala la reconoce desde la siguiente lectura sin que nadie tenga
        que darla de alta dos veces.
        """
        perfil = await self.obtener_perfil(db, usuario_id)

        if vehiculo_id:
            vehiculo = await db.get(Vehiculo, vehiculo_id)
            if not vehiculo:
                raise EntidadNoEncontrada("Vehículo no encontrado")
            if vehiculo.socio_id and vehiculo.socio_id != usuario_id:
                dueno = await db.get(Usuario, vehiculo.socio_id)
                raise EntidadDuplicada(
                    f"La placa {vehiculo.placa} ya está a nombre de "
                    f"{dueno.nombre_completo if dueno else 'otra persona'}. "
                    f"El cambio de titular se hace desde Parque Automotor."
                )
            vehiculo.socio_id = usuario_id
        else:
            datos_nuevos = datos_nuevos or {}
            placa = normalizar_placa(datos_nuevos.get("placa", ""))
            if not placa:
                raise EntidadNoEncontrada("Hace falta la placa")

            res = await db.execute(select(Vehiculo).where(Vehiculo.placa == placa))
            existente = res.scalars().first()
            if existente:
                if existente.socio_id and existente.socio_id != usuario_id:
                    dueno = await db.get(Usuario, existente.socio_id)
                    raise EntidadDuplicada(
                        f"La placa {placa} ya está a nombre de "
                        f"{dueno.nombre_completo if dueno else 'otra persona'}."
                    )
                existente.socio_id = usuario_id
                vehiculo = existente
            else:
                vehiculo = Vehiculo(
                    placa=placa,
                    marca=(datos_nuevos.get("marca") or "SIN ESPECIFICAR").upper(),
                    modelo=(datos_nuevos.get("modelo") or "SIN ESPECIFICAR").upper(),
                    color=(datos_nuevos.get("color") or "SIN ESPECIFICAR").upper(),
                    tipo=datos_nuevos.get("tipo"),
                    socio_id=usuario_id,
                )
                db.add(vehiculo)

        # Tener una placa a su nombre convierte la declaración en un hecho.
        perfil.tiene_vehiculo = True

        await db.commit()
        await db.refresh(vehiculo)
        return vehiculo

    # ─── Exportaciones ────────────────────────────────────────────────────────

    async def exportar_excel(
        self, db: AsyncSession, dormitorio_id: UUID
    ) -> tuple[io.BytesIO, str]:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        dormitorio = await self.obtener_dormitorio(db, dormitorio_id)

        res = await db.execute(
            select(Habitacion)
            .where(
                Habitacion.dormitorio_id == dormitorio_id,
                Habitacion.activo == True,  # noqa: E712
            )
            .order_by(
                func.coalesce(func.length(Habitacion.piso), 0),
                Habitacion.piso.asc().nulls_first(),
                func.length(Habitacion.numero),
                Habitacion.numero,
            )
        )
        habitaciones = res.scalars().all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Dormitorio"

        titulo_font = Font(bold=True, color="FFFFFF", size=11)
        titulo_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
        hab_font = Font(bold=True, size=11)
        hab_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        centrado = Alignment(horizontal="center", vertical="center")
        borde_fino = Border(bottom=Side(style="thin", color="CCCCCC"))

        columnas = [
            "Habitación", "Piso", "Camas", "Grado", "Nombre", "Apellido",
            "Cédula", "Teléfono", "Unidad", "Jefe directo", "Tel. jefe",
            "Vehículos",
        ]
        for col, titulo in enumerate(columnas, 1):
            celda = ws.cell(row=1, column=col, value=titulo)
            celda.font = titulo_font
            celda.fill = titulo_fill
            celda.alignment = centrado

        fila = 2
        for h in habitaciones:
            integrantes = await self.listar_integrantes_habitacion(db, h.id)
            ocupacion = len(integrantes)
            piso_txt = f" · Piso {h.piso}" if h.piso else ""
            resumen = f"Habitación {h.numero}{piso_txt} — {ocupacion}/{h.camas} camas"

            ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=len(columnas))
            celda_hab = ws.cell(row=fila, column=1, value=resumen)
            celda_hab.font = hab_font
            celda_hab.fill = hab_fill
            fila += 1

            if not integrantes:
                ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=len(columnas))
                ws.cell(row=fila, column=1, value="(vacía)").font = Font(italic=True, color="888888")
                fila += 1
            else:
                for inte in integrantes:
                    placas = ", ".join(v.placa for v in inte.vehiculos) if inte.vehiculos else "—"
                    ws.cell(row=fila, column=1, value=h.numero).alignment = centrado
                    ws.cell(row=fila, column=2, value=h.piso or "—").alignment = centrado
                    ws.cell(row=fila, column=3, value=h.camas).alignment = centrado
                    ws.cell(row=fila, column=4, value=inte.grado or "—")
                    ws.cell(row=fila, column=5, value=inte.nombre)
                    ws.cell(row=fila, column=6, value=inte.apellido)
                    ws.cell(row=fila, column=7, value=inte.cedula)
                    ws.cell(row=fila, column=8, value=inte.telefono or "—")
                    ws.cell(row=fila, column=9, value=inte.unidad or "—")
                    ws.cell(row=fila, column=10, value=inte.jefe_nombre or "—")
                    ws.cell(row=fila, column=11, value=inte.jefe_telefono or "—")
                    ws.cell(row=fila, column=12, value=placas)
                    for c in range(1, len(columnas) + 1):
                        ws.cell(row=fila, column=c).border = borde_fino
                    fila += 1

        anchos = [12, 8, 8, 14, 18, 18, 14, 14, 22, 22, 14, 16]
        for i, ancho in enumerate(anchos, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = ancho

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        nombre = dormitorio.nombre.replace(" ", "_")
        return buf, f"{nombre}_personal.xlsx"

    async def exportar_qr_zip(
        self, db: AsyncSession, dormitorio_id: UUID, frontend_url: str
    ) -> tuple[io.BytesIO, str, int]:
        import zipfile

        import qrcode
        from PIL import Image, ImageDraw, ImageFont

        dormitorio = await self.obtener_dormitorio(db, dormitorio_id)

        res = await db.execute(
            select(Habitacion)
            .where(
                Habitacion.dormitorio_id == dormitorio_id,
                Habitacion.activo == True,  # noqa: E712
            )
            .order_by(
                func.coalesce(func.length(Habitacion.piso), 0),
                Habitacion.piso.asc().nulls_first(),
                func.length(Habitacion.numero),
                Habitacion.numero,
            )
        )
        habitaciones = res.scalars().all()

        def _cargar_fuente(tam: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
            for nombre_fuente in (
                "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
                "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                "/usr/share/fonts/TTF/DejaVuSans-Bold.ttf",
                "arial.ttf",
                "Arial.ttf",
            ):
                try:
                    return ImageFont.truetype(nombre_fuente, tam)
                except OSError:
                    continue
            try:
                return ImageFont.load_default(size=tam)
            except TypeError:
                return ImageFont.load_default()

        fuente = _cargar_fuente(28)
        fuente_peq = _cargar_fuente(18)

        sin_token = 0
        buf_zip = io.BytesIO()

        with zipfile.ZipFile(buf_zip, "w", zipfile.ZIP_DEFLATED) as zf:
            for h in habitaciones:
                if not h.token:
                    sin_token += 1
                    continue

                url = f"{frontend_url.rstrip('/')}/habitacion/{h.token}"

                qr = qrcode.QRCode(
                    box_size=12, border=2,
                    error_correction=qrcode.constants.ERROR_CORRECT_H,
                )
                qr.add_data(url)
                qr.make(fit=True)
                qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
                qr_w, qr_h = qr_img.size

                margen = 40
                alto_etiqueta = 80
                ancho_total = qr_w + margen * 2
                alto_total = qr_h + margen * 2 + alto_etiqueta

                canvas = Image.new("RGB", (ancho_total, alto_total), "white")
                canvas.paste(qr_img, (margen, margen))

                draw = ImageDraw.Draw(canvas)

                linea1 = f"Habitacion {h.numero}"
                if h.piso:
                    linea1 += f"  -  Piso {h.piso}"
                linea2 = dormitorio.nombre

                bbox1 = draw.textbbox((0, 0), linea1, font=fuente)
                x1 = (ancho_total - (bbox1[2] - bbox1[0])) // 2
                y1 = margen + qr_h + 10
                draw.text((x1, y1), linea1, fill="black", font=fuente)

                bbox2 = draw.textbbox((0, 0), linea2, font=fuente_peq)
                x2 = (ancho_total - (bbox2[2] - bbox2[0])) // 2
                y2 = y1 + (bbox1[3] - bbox1[1]) + 6
                draw.text((x2, y2), linea2, fill="#555555", font=fuente_peq)

                img_buf = io.BytesIO()
                canvas.save(img_buf, format="JPEG", quality=92)
                img_buf.seek(0)

                piso_parte = f"_Piso_{h.piso}" if h.piso else ""
                nombre_archivo = f"Hab_{h.numero}{piso_parte}.jpg"
                zf.writestr(nombre_archivo, img_buf.getvalue())

        buf_zip.seek(0)
        nombre = dormitorio.nombre.replace(" ", "_")
        return buf_zip, f"{nombre}_QR_puertas.zip", sin_token


dormitorio_service = DormitorioService()
